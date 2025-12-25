"""Скрипт для сбора артикулов товаров Ozon и записи в Articles.xlsx."""
import sys
from pathlib import Path

# Добавляем src в путь
sys.path.insert(0, str(Path(__file__).parent / "src"))

from loguru import logger
from config.settings import Settings
from utils.logger import setup_logger
from api.ozon_api import OzonAPI
import openpyxl
from openpyxl import load_workbook


def collect_ozon_articles(settings: Settings, articles_file_path: Path) -> bool:
    """Собрать артикулы товаров Ozon и записать в Articles.xlsx.
    
    Args:
        settings: Настройки приложения
        articles_file_path: Путь к файлу Articles.xlsx
        
    Returns:
        True если успешно, False иначе
    """
    logger.info("=" * 60)
    logger.info("СБОР АРТИКУЛОВ ТОВАРОВ OZON")
    logger.info("=" * 60)
    
    ozon_api_keys = settings.get_ozon_api_keys()
    ozon_client_ids = settings.get_ozon_client_ids()
    
    all_articles = {}  # {cabinet_name: [articles]}
    
    # Собираем артикулы из всех кабинетов Ozon
    for cabinet_name, api_key in ozon_api_keys.items():
        if not api_key:
            logger.warning(f"Пропускаем кабинет {cabinet_name} - нет API ключа")
            continue
        
        client_id = ozon_client_ids.get(cabinet_name)
        if not client_id:
            logger.warning(f"Пропускаем кабинет {cabinet_name} - нет Client ID")
            continue
        
        logger.info(f"Обработка кабинета: {cabinet_name}")
        
        try:
            api = OzonAPI(api_key, client_id, request_delay=0.5)
            
            # Получаем все товары из кабинета
            logger.info("Получение списка товаров...")
            products = api.get_all_products()
            
            if not products:
                logger.warning(f"Не удалось получить товары для кабинета {cabinet_name}")
                continue
            
            # Извлекаем артикулы (offer_id)
            articles = []
            for product in products:
                offer_id = product.get("offer_id") or product.get("offerId") or product.get("offer_id")
                if offer_id:
                    articles.append(str(offer_id))
                else:
                    # Если нет offer_id, пробуем другие поля
                    product_id = product.get("product_id") or product.get("productId")
                    if product_id:
                        logger.debug(f"Товар без offer_id, используем product_id: {product_id}")
            
            all_articles[cabinet_name] = articles
            logger.success(f"Собрано артикулов из кабинета {cabinet_name}: {len(articles)}")
            
        except Exception as e:
            logger.error(f"Ошибка при обработке кабинета {cabinet_name}: {e}")
            logger.exception("Детали ошибки:")
            continue
    
    if not all_articles:
        logger.error("Не удалось собрать артикулы ни из одного кабинета")
        return False
    
    # Записываем в Articles.xlsx
    logger.info("")
    logger.info("Запись артикулов в Articles.xlsx...")
    
    try:
        # Загружаем существующий файл или создаём новый
        if articles_file_path.exists():
            wb = load_workbook(articles_file_path)
            logger.info(f"Файл {articles_file_path} загружен")
        else:
            wb = openpyxl.Workbook()
            logger.info(f"Создан новый файл {articles_file_path}")
        
        # Создаём или обновляем лист для Ozon
        sheet_name = "OzonArticles"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"Лист {sheet_name} уже существует, будет обновлён")
            # Очищаем существующие данные (кроме заголовка)
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)
        else:
            ws = wb.create_sheet(sheet_name)
            logger.info(f"Создан новый лист {sheet_name}")
        
        # Записываем заголовки
        ws['A1'] = "Кабинет"
        ws['B1'] = "Артикул (offer_id)"
        
        # Записываем данные
        row = 2
        for cabinet_name, articles in all_articles.items():
            for article in articles:
                ws[f'A{row}'] = cabinet_name
                ws[f'B{row}'] = article
                row += 1
        
        # Сохраняем файл
        wb.save(articles_file_path)
        logger.success(f"Артикулы записаны в {articles_file_path} на лист '{sheet_name}'")
        logger.info(f"Всего записано артикулов: {sum(len(arts) for arts in all_articles.values())}")
        
        return True
        
    except Exception as e:
        logger.error(f"Ошибка при записи в Articles.xlsx: {e}")
        logger.exception("Детали ошибки:")
        return False


def main():
    """Основная функция."""
    try:
        # Загрузка настроек
        settings = Settings()
        
        # Настройка логирования
        setup_logger(settings.logs_dir, debug=settings.debug)
        
        logger.info("=" * 60)
        logger.info("ЗАПУСК СБОРА АРТИКУЛОВ OZON")
        logger.info("=" * 60)
        
        # Путь к Articles.xlsx
        # Ищем файл в разных возможных местах
        possible_paths = [
            Path("D:/Zhenya/Bullshit/WBagentforDasha/.cursor/Articles.xlsx"),
            Path("../WBagentforDasha/.cursor/Articles.xlsx"),
            Path("Articles.xlsx"),
        ]
        
        articles_file_path = None
        for path in possible_paths:
            if path.exists():
                articles_file_path = path
                break
        
        if not articles_file_path:
            logger.error("Файл Articles.xlsx не найден")
            logger.info("Проверьте наличие файла в одном из мест:")
            for path in possible_paths:
                logger.info(f"  - {path}")
            return 1
        
        logger.info(f"Используется файл: {articles_file_path}")
        
        # Собираем артикулы
        success = collect_ozon_articles(settings, articles_file_path)
        
        if success:
            logger.success("=" * 60)
            logger.success("СБОР АРТИКУЛОВ ЗАВЕРШЁН УСПЕШНО")
            logger.success("=" * 60)
            return 0
        else:
            logger.error("=" * 60)
            logger.error("ОШИБКА ПРИ СБОРЕ АРТИКУЛОВ")
            logger.error("=" * 60)
            return 1
        
    except KeyboardInterrupt:
        logger.warning("Прервано пользователем")
        return 1
        
    except Exception as e:
        logger.error(f"Критическая ошибка: {e}")
        logger.exception("Детали ошибки:")
        return 1


if __name__ == "__main__":
    sys.exit(main())
