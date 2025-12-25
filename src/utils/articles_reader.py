"""Утилита для чтения артикулов из Articles.xlsx."""
from pathlib import Path
from typing import List, Dict, Optional
from loguru import logger
import openpyxl
from openpyxl import load_workbook


def read_wb_articles(articles_file_path: Path) -> List[str]:
    """Читать артикулы Wildberries из Articles.xlsx.
    
    Args:
        articles_file_path: Путь к файлу Articles.xlsx
        
    Returns:
        Список артикулов WB
    """
    articles = []
    
    if not articles_file_path.exists():
        logger.error(f"Файл {articles_file_path} не найден")
        return articles
    
    try:
        wb = load_workbook(articles_file_path, read_only=True, data_only=True)
        
        # Ищем лист с артикулами WB
        # Возможные названия листов: "WBarticules", "WB", "Wildberries", и т.д.
        sheet_names = wb.sheetnames
        wb_sheet = None
        
        for sheet_name in sheet_names:
            if any(keyword in sheet_name.lower() for keyword in ['wb', 'wildberries', 'article']):
                wb_sheet = wb[sheet_name]
                logger.info(f"Найден лист с артикулами WB: {sheet_name}")
                break
        
        if not wb_sheet:
            # Если не нашли, пробуем первый лист
            wb_sheet = wb[wb.sheetnames[0]]
            logger.warning(f"Лист с артикулами WB не найден, используем первый лист: {wb_sheet.title}")
        
        # Читаем артикулы (предполагаем, что они в одной из колонок)
        # Обычно артикулы в колонке B или C
        for row in wb_sheet.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
            # Пробуем разные колонки
            for cell_value in row:
                if cell_value:
                    article = str(cell_value).strip()
                    # Проверяем, что это похоже на артикул (не пусто, не заголовок)
                    if article and article.lower() not in ['артикул', 'article', 'vendorcode', 'vendor_code']:
                        articles.append(article)
                        break
        
        wb.close()
        logger.success(f"Прочитано артикулов WB: {len(articles)}")
        
    except Exception as e:
        logger.error(f"Ошибка при чтении Articles.xlsx: {e}")
        logger.exception("Детали ошибки:")
    
    return articles


def find_articles_file() -> Optional[Path]:
    """Найти файл Articles.xlsx в возможных местах.
    
    Returns:
        Путь к файлу или None
    """
    possible_paths = [
        Path("D:/Zhenya/Bullshit/WBagentforDasha/.cursor/Articles.xlsx"),
        Path("../WBagentforDasha/.cursor/Articles.xlsx"),
        Path("../../WBagentforDasha/.cursor/Articles.xlsx"),
        Path("Articles.xlsx"),
        Path("./Articles.xlsx"),
    ]
    
    for path in possible_paths:
        if path.exists():
            logger.info(f"Найден файл Articles.xlsx: {path}")
            return path
    
    logger.warning("Файл Articles.xlsx не найден ни в одном из возможных мест")
    return None

