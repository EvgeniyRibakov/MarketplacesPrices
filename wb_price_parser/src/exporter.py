"""
Экспорт данных в Excel файл (.xlsx)
"""
import logging
from pathlib import Path
from typing import List, Optional
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .models import Product
from .config import Config


class ExcelExporter:
    """Класс для экспорта данных в Excel"""
    
    def __init__(self, config: Config, logger: Optional[logging.Logger] = None):
        """
        Инициализация экспортера
        
        Args:
            config: Конфигурация проекта
            logger: Логгер (опционально)
        """
        self.config = config
        self.logger = logger or logging.getLogger(__name__)
    
    def export(self, products: List[Product], output_file: Optional[str] = None) -> Path:
        """
        Экспорт списка товаров в Excel файл
        
        Args:
            products: Список товаров для экспорта
            output_file: Имя файла (если не указано, берётся из конфига)
        
        Returns:
            Путь к созданному файлу
        """
        if not products:
            self.logger.warning("Список товаров пуст, файл не будет создан")
            return None
        
        # Определяем имя файла
        if output_file:
            file_path = Path(output_file)
        else:
            file_path = self.config.output_dir / self.config.output_file
        
        # Создаём директорию если нужно
        file_path.parent.mkdir(parents=True, exist_ok=True)
        
        self.logger.info(f"Экспорт {len(products)} товаров в {file_path}")
        
        # Конвертируем товары в словари
        data = [product.to_dict() for product in products]
        
        # Создаём DataFrame
        df = pd.DataFrame(data)
        
        # Определяем порядок колонок согласно ТЗ
        columns_order = [
            "cabinet_name",
            "cabinet_id",
            "article",
            "name",
            "price_basic",
            "price_product",
            "price_card",
            "source_price_basic",
            "source_price_product",
            "source_price_card",
            "product_url",
            "timestamp",
        ]
        
        # Упорядочиваем колонки (только те, что есть в данных)
        existing_columns = [col for col in columns_order if col in df.columns]
        df = df[existing_columns]
        
        # Сохраняем в Excel
        df.to_excel(file_path, index=False, engine="openpyxl")
        
        # Форматируем файл
        self._format_excel(file_path)
        
        self.logger.info(f"Экспорт завершён: {file_path}")
        return file_path
    
    def _format_excel(self, file_path: Path):
        """
        Форматирование Excel файла (заголовки, стили)
        
        Args:
            file_path: Путь к Excel файлу
        """
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Стиль для заголовков
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Форматируем заголовки
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Автоподбор ширины колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Устанавливаем ширину с небольшим запасом
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Форматируем числовые колонки (цены)
            price_columns = ["E", "F", "G"]  # price_basic, price_product, price_card
            for col in price_columns:
                for row in range(2, ws.max_row + 1):
                    cell = ws[f"{col}{row}"]
                    if cell.value is not None:
                        cell.number_format = "#,##0.00"
            
            # Сохраняем изменения
            wb.save(file_path)
            
        except Exception as e:
            self.logger.warning(f"Не удалось отформатировать Excel файл: {e}")

